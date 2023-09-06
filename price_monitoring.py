import pandas as pd
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime

# mydataset = {
#         'cars' : ["BMW", "Volvo", "Ford"],
#         'passings' : [3, 7, 2]
#         }
# myvar = pd.DataFrame(mydataset)
#
# diet = {
#         'calories': ["420", "380", "400"],
#         'duration': ["50", "45", "40"]
# }
# mydiet = pd.DataFrame(diet)
# df = pd.read_csv('file.csv')
# df2 = df.to_dict()

book = openpyxl.open('file.xlsx', read_only=True)

sheet = book.active
first_link = sheet[2][0].value
first_min_price = sheet[2][1].value
second_max_price = sheet[2][2].value


# for row in range(1, sheet.max_row + 1):
#     link = sheet[row][0].value
#     min_price = sheet[row][1].value
#     max_price = sheet[row][2].value
#     print(row, link, min_price, max_price)
class Monitoring(webdriver.Chrome):
    def __init__(self, driver_path=r"C:/SeleniumDrivers/chromedriver", teardown=False):
        self.teardown = teardown
        self.driver_path = driver_path
        os.environ['PATH'] += self.driver_path
        super(Monitoring, self).__init__()
        self.implicitly_wait(20)
        self.maximize_window()

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()

    def land_first_page(self):
        self.get(first_link)

    def filtres(self):
        min_price_input = self.find_element(By.ID, 'minPrice_')
        min_price_input.send_keys(first_min_price)
        max_price_input = self.find_element(By.ID, 'maxPrice_')
        max_price_input.send_keys(sheet[2][2].value)
        time.sleep(3)
        search = self.find_element(By.CSS_SELECTOR,
                                   'a[href="https://m.ua/ua/m1_magazilla.php?katalog_=206&pf_=1&minPrice_=10000&maxPrice_=20000&sc_id_=980&order_=pop&save_podbor_=1"]')
        search.click()
        filtres = self.find_element(By.ID, 'order_label')
        filtres.click()
        time.sleep(1)
        from_cheapest = self.find_element(By.XPATH,
                                          '/html/body/div[1]/table[3]/tbody/tr/td[3]/table[1]/tbody/tr/td[1]/div[2]/div[1]/a[2]')
        from_cheapest.click()

    def create_book(self):
        book = openpyxl.Workbook()

        sheet = book.active
        sheet['A2'] = 242
        sheet['B3'] = 'hello'

        sheet[1][0].value = 'world'
        sheet.cell(row=1, column=1).value = 'HelloBabies'

        book.save('my_book.xlsx')
        book.close()

    def take_data(self):
        results = openpyxl.Workbook()
        sheet1 = results.active
        first_cell = 1
        block = self.find_element(By.XPATH, '//*[@id="list"]/tbody')
        blocks = block.find_elements(By.CLASS_NAME, 'list-tr')

        for camera_title in blocks:
            all_cameras_titles = camera_title.find_elements(By.CLASS_NAME, 'list-model-title')

            for camera_href in all_cameras_titles:
                href_blocks = camera_href.find_elements(By.TAG_NAME, 'a')

                for href in href_blocks:
                    one_href = href.get_attribute('href')
                    sheet1.cell(row=first_cell, column=1).value = datetime.date.today()
                    sheet1.cell(row=first_cell, column=2).value = first_cell
                    sheet1.cell(row=first_cell, column=3).value = one_href
                    results.save('results.xlsx')
                    results.close()
                    first_cell += 1

        first_cell = 1
        for camera_lowest_price in blocks:
            all_cameras_lowest_prices = camera_lowest_price.find_elements(By.CLASS_NAME, 'list-big-price')

            for camera_lowest_prices in all_cameras_lowest_prices:
                lowest_prices_upper = camera_lowest_prices.find_elements(By.CSS_SELECTOR, 'a[title="Порівняти ціни! "]')

                for camera_lowest_price_spans in lowest_prices_upper:
                    lowest_price_blocks = camera_lowest_price_spans.find_elements(By.TAG_NAME, 'span')

                    for lowest_price in lowest_price_blocks:
                        one_span = lowest_price.get_attribute('textContent')
                        sheet1.cell(row=first_cell, column=4).value = one_span
                        results.save('results.xlsx')
                        results.close()
                        first_cell +=1


        first_cell = 1
        for cameras_description in blocks:
            all_cameras_description = cameras_description.find_elements(By.CLASS_NAME, 'list-model-desc')

            for description in all_cameras_description:
                one_description = description.get_attribute('textContent')
                sheet1.cell(row=first_cell, column=5).value = one_description
                results.save('results.xlsx')
                results.close()
                first_cell += 1

